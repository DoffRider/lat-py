import openpyxl

inv = openpyxl.load_workbook("inventory.xlsx")
# nama sheetnya
produk_list_sheet = inv["Sheet1"]

# dictionary kosong
produk_dict = {}
total_value_per_sup_dict = {}
produk_under_10 = {}

# looping sesuai dengan berapa banyak baris
# dimulai dari baris kedua
# ditambah satu karena nomer index dimulai dari 0
for baris_produk in range(2, produk_list_sheet.max_row + 1):
    # mencari baris dan kolom dari nama suplayer
    # pertama baris(samping) kedua kolom (bawah)
    # baris sudah dinamis mengikuti file dengan looping
    # ditambah .value dibelakang agar membaca isi cellnya jika tidak ada maka hanya akan muncul info objeknya saja seperti (A4, B4)
    nama_suplayer = produk_list_sheet.cell(baris_produk,4).value

    inventory = produk_list_sheet.cell(baris_produk, 2).value
    price = produk_list_sheet.cell(baris_produk, 3).value
    produk_num = produk_list_sheet.cell(baris_produk,1).value
    # karena ingin membuat/update kolom/baris tidak perlu di tambah .value karena kita perlu objeknya untuk diolah
    inventory_price = produk_list_sheet.cell(baris_produk,5)
 

    if nama_suplayer in produk_dict:
        produk_dict[nama_suplayer] = produk_dict[nama_suplayer] + 1
        # print(produk_dict[nama_suplayer])

    else:
        print(f"suplayer {nama_suplayer} baru ditambahkan!")
        produk_dict[nama_suplayer] = 1


# perhitungan total value dari invetory per suplay

    if nama_suplayer in total_value_per_sup_dict:
        total_value_per_sup_dict[nama_suplayer] = total_value_per_sup_dict[nama_suplayer]+inventory*price
    else:
        total_value_per_sup_dict[nama_suplayer] = inventory*price

#logika produk dengan inventori kurang dari 10
    if inventory < 10 :
        produk_under_10[int(produk_num)] = int(inventory)

# menambah value untuk total inventory price
    inventory_price.value = inventory*price


print("\n",produk_dict)
print("\n",total_value_per_sup_dict)
print("\n",produk_under_10)

# save file baru
inv.save("inventory_baru.xlsx")